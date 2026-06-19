"""
AI app views — API endpoints for the chat popup.

Persistent chat sessions:
  POST /ai/api/sessions/new/        — create a new session, returns session_id + title
  GET  /ai/api/sessions/            — list all sessions for the current user
  GET  /ai/api/session/<id>/        — fetch all messages for a session
  POST /ai/api/chat/                — send a message (pass session_id in body)
  POST /ai/api/clear/               — clear in-memory context (session_id in body)

File tools:
  GET  /ai/api/files/               — list workspace files
  GET  /ai/api/download/<filename>/ — download a file
  GET  /ai/api/read/<filename>/     — read file content
  POST /ai/api/save/<filename>/     — save file content
  POST /ai/api/sheet-edit/          — structured spreadsheet operations
  GET  /ai/sheet-editor/<filename>/ — render sheet editor page
"""

import json
import os
from pathlib import Path

from django.http import JsonResponse, FileResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.conf import settings
from django.shortcuts import render

from .gemini_service import send_message, generate_sheet_ops
from .ai_tools import _workspace_root, _safe_path, list_files


MAX_CHAT_BODY_BYTES = int(os.environ.get('AI_MAX_CHAT_BODY_BYTES', 12000))


def _auth_required_response():
    return JsonResponse(
        {
            'status': 'error',
            'response': 'Please login or sign up to use SkipStep AI.',
            'login_url': '/login/',
            'signup_url': '/register/',
        },
        status=403,
    )


def _get_session_id(request):
    """Derive a stable session id from Django's session framework."""
    if not request.session.session_key:
        request.session.create()
    return request.session.session_key


def _get_workspace_id(request):
    """Return a consistent workspace ID: user ID if logged in, else session key."""
    if request.user.is_authenticated:
        return f'user_{request.user.id}'
    return _get_session_id(request)


# ---------------------------------------------------------------------------
# Persistent Chat Session endpoints
# ---------------------------------------------------------------------------

@csrf_exempt
@require_POST
def new_chat_session_api(request):
    """Create a new AI chat session in the database and return its ID."""
    if not request.user.is_authenticated:
        return _auth_required_response()

    from .models import AIChatSession
    session = AIChatSession.objects.create(user=request.user, title='New Chat')
    return JsonResponse({
        'status': 'ok',
        'session_id': session.id,
        'title': session.title,
    })


@require_GET
def list_chat_sessions_api(request):
    """Return all chat sessions for the logged-in user, newest first."""
    if not request.user.is_authenticated:
        return _auth_required_response()

    from .models import AIChatSession
    sessions = AIChatSession.objects.filter(user=request.user).order_by('-updated_at')[:50]
    return JsonResponse({
        'status': 'ok',
        'sessions': [
            {
                'id': s.id,
                'title': s.title,
                'created_at': s.created_at.isoformat(),
                'updated_at': s.updated_at.isoformat(),
            }
            for s in sessions
        ],
    })


@require_GET
def get_chat_session_api(request, session_id):
    """Return all messages for a specific chat session."""
    if not request.user.is_authenticated:
        return _auth_required_response()

    from .models import AIChatSession
    try:
        session = AIChatSession.objects.get(id=session_id, user=request.user)
    except AIChatSession.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Session not found.'}, status=404)

    messages = session.messages.order_by('created_at')
    return JsonResponse({
        'status': 'ok',
        'session_id': session.id,
        'title': session.title,
        'messages': [
            {
                'role': m.role,
                'content': m.content,
                'file_actions': m.file_actions,
                'created_at': m.created_at.isoformat(),
            }
            for m in messages
        ],
    })


@csrf_exempt
@require_POST
def delete_chat_session_api(request, session_id):
    """Delete a specific chat session and all its messages."""
    if not request.user.is_authenticated:
        return _auth_required_response()

    from .models import AIChatSession
    try:
        session = AIChatSession.objects.get(id=session_id, user=request.user)
    except AIChatSession.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Session not found.'}, status=404)

    session.delete()
    return JsonResponse({'status': 'ok', 'message': 'Session deleted.'})


# ---------------------------------------------------------------------------
# Chat API
# ---------------------------------------------------------------------------

@csrf_exempt
@require_POST
def chat_api(request):
    """
    Receive a user message and return the AI response.
    Expects JSON body: { "message": "...", "session_id": <int or null> }
    If session_id is provided and belongs to the user, saves messages to DB.
    """
    if not request.user.is_authenticated:
        return _auth_required_response()

    if len(request.body) > MAX_CHAT_BODY_BYTES:
        return JsonResponse({'status': 'error', 'response': 'Request is too large.'}, status=413)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'response': 'Invalid JSON.'}, status=400)

    user_message = body.get('message', '').strip()
    if not user_message:
        return JsonResponse({'status': 'error', 'response': 'Empty message.'}, status=400)
        
    active_file = body.get('active_file')
    prompt_text = user_message
    if active_file:
        prompt_text = f"[System Context: The user is currently viewing/editing the file '{active_file}' in their active tab. If they ask to modify 'this file' or add something to 'the document/sheet', assume they mean '{active_file}'.]\n\n{user_message}"

    db_session_id = body.get('session_id')  # Optional: persist to DB
    db_session = None

    # Resolve the DB session if ID was provided
    if db_session_id:
        from .models import AIChatSession
        try:
            db_session = AIChatSession.objects.get(id=db_session_id, user=request.user)
        except AIChatSession.DoesNotExist:
            db_session = None

    # Prepare stateless history array
    messages_data = []
    if db_session:
        db_messages = db_session.messages.order_by('created_at').values('role', 'content')
        messages_data = list(db_messages)
    else:
        messages_data = request.session.get('ai_chat_history', [])

    workspace_id = _get_workspace_id(request)
    
    from asgiref.sync import async_to_sync
    result = async_to_sync(send_message)(messages_data, workspace_id, prompt_text)

    # Persist the new turns
    if result.get('status') == 'ok':
        if db_session:
            from .models import AIChatMessage
            AIChatMessage.objects.create(session=db_session, role='user', content=user_message, file_actions=[])
            AIChatMessage.objects.create(session=db_session, role='model', content=result.get('response', ''), file_actions=result.get('file_actions', []))

            if db_session.title == 'New Chat':
                db_session.title = user_message[:60].strip()
                db_session.save(update_fields=['title', 'updated_at'])
            else:
                db_session.save(update_fields=['updated_at'])
        else:
            messages_data.append({'role': 'user', 'content': user_message})
            messages_data.append({'role': 'model', 'content': result.get('response', '')})
            request.session['ai_chat_history'] = messages_data[-40:]
            request.session.modified = True

    return JsonResponse(result)


@csrf_exempt
@require_POST
def clear_chat_api(request):
    """
    Clear the current guest session's memory.
    """
    if 'ai_chat_history' in request.session:
        del request.session['ai_chat_history']
        request.session.modified = True

    return JsonResponse({'status': 'ok', 'message': 'Chat context cleared.'})


@csrf_exempt
@require_POST
def switch_model_api(request):
    """Switch the Gemini model being used for AI interactions."""
    if not request.user.is_authenticated:
        return _auth_required_response()

    try:
        body = json.loads(request.body)
        model = body.get('model', '').strip()
    except Exception:
        return JsonResponse({'status': 'error', 'response': 'Invalid data.'}, status=400)

    if model not in ('gemini-2.5-flash', 'gemini-1.5-pro', 'gemini-2.5-pro'):
        return JsonResponse({'status': 'error', 'response': 'Invalid model selection.'}, status=400)

    # Update .env
    try:
        from pathlib import Path
        from django.conf import settings
        import re
        import os
        
        env_path = Path(settings.BASE_DIR) / '.env'
        if env_path.exists():
            content = env_path.read_text(encoding='utf-8')
            if 'GEMINI_MODEL=' in content:
                content = re.sub(r'GEMINI_MODEL=.*', f'GEMINI_MODEL={model}', content)
            else:
                content += f'\nGEMINI_MODEL={model}\n'
            env_path.write_text(content, encoding='utf-8')
        
        # Update running OS environment and module variable dynamically
        os.environ['GEMINI_MODEL'] = model
        from . import gemini_service
        gemini_service.GEMINI_MODEL = model

        return JsonResponse({'status': 'ok', 'response': f'Model switched to {model}.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'response': str(e)}, status=500)


# ---------------------------------------------------------------------------
# File management endpoints
# ---------------------------------------------------------------------------

@require_GET
def list_files_api(request):
    """Return a list of files in the user's AI workspace."""
    if not request.user.is_authenticated:
        return _auth_required_response()

    workspace_id = _get_workspace_id(request)
    result = list_files(workspace_id)
    return JsonResponse(result)


@require_GET
def download_file(request, filename):
    """Download a file from the AI workspace."""
    if not request.user.is_authenticated:
        raise Http404('Login required.')

    workspace_id = _get_workspace_id(request)
    try:
        path = _safe_path(workspace_id, filename)
    except ValueError:
        raise Http404('Access denied.')

    if not path.exists():
        raise Http404('File not found.')

    return FileResponse(open(path, 'rb'), as_attachment=True, filename=filename)


@require_GET
def read_file_api(request, filename):
    """Read a file from the AI workspace for the frontend editor."""
    if not request.user.is_authenticated:
        return _auth_required_response()

    workspace_id = _get_workspace_id(request)
    try:
        path = _safe_path(workspace_id, filename)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Access denied.'}, status=403)

    if not path.exists():
        return JsonResponse({'status': 'error', 'message': 'File not found.'}, status=404)

    try:
        if path.name.lower().endswith('.xlsx'):
            import openpyxl
            import io
            import csv
            import zipfile
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                output = io.StringIO()
                writer = csv.writer(output)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([str(cell) if cell is not None else "" for cell in row])
                content = output.getvalue()
            except zipfile.BadZipFile:
                content = path.read_text(encoding='utf-8')
        else:
            content = path.read_text(encoding='utf-8')

        meta_content = None
        meta_path = path.parent / f"{path.name}.meta.json"
        if meta_path.exists():
            meta_content = meta_path.read_text(encoding='utf-8')

        return JsonResponse({'status': 'ok', 'content': content, 'meta': meta_content})
    except UnicodeDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Cannot edit binary files.'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error reading file: {str(e)}'}, status=500)


@csrf_exempt
@require_POST
def save_file_api(request, filename):
    """Save user edits back to a file in the AI workspace."""
    if not request.user.is_authenticated:
        return _auth_required_response()

    workspace_id = _get_workspace_id(request)
    try:
        path = _safe_path(workspace_id, filename)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Access denied.'}, status=403)

    if not path.exists():
        return JsonResponse({'status': 'error', 'message': 'File not found.'}, status=404)

    try:
        body = json.loads(request.body)
        content = body.get('content')
        if content is not None:
            if path.name.lower().endswith('.xlsx'):
                import openpyxl
                import csv
                import io
                wb = openpyxl.Workbook()
                ws = wb.active
                reader = csv.reader(io.StringIO(content))
                for row in reader:
                    ws.append(row)
                wb.save(path)
            else:
                path.write_text(content, encoding='utf-8')

        meta_content = body.get('meta')
        if meta_content is not None:
            meta_path = path.parent / f"{path.name}.meta.json"
            meta_path.write_text(meta_content, encoding='utf-8')

        return JsonResponse({'status': 'ok', 'message': 'File saved.'})
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON.'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def sheet_editor_page(request, filename):
    """Render the live sheet editor page for a CSV file."""
    if not request.user.is_authenticated:
        from django.shortcuts import redirect
        return redirect('/login/')
    return render(request, 'user/sheet_editor.html', {'filename': filename})


@csrf_exempt
@require_POST
def sheet_edit_api(request):
    """Process a natural language instruction against a spreadsheet via Gemini structured output."""
    if not request.user.is_authenticated:
        return _auth_required_response()

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON.'}, status=400)

    sheet_data = body.get('sheet_data', [])
    prompt = body.get('prompt', '').strip()
    filename = body.get('filename', '')

    if not prompt:
        return JsonResponse({'status': 'error', 'message': 'Empty instruction.'}, status=400)
    if not sheet_data:
        return JsonResponse({'status': 'error', 'message': 'No sheet data provided.'}, status=400)

    from asgiref.sync import async_to_sync
    result = async_to_sync(generate_sheet_ops)(sheet_data, prompt)

    if result['status'] != 'ok':
        return JsonResponse(result, status=500)

    # Save updated sheet back to file
    save_warning = None
    if filename:
        workspace_id = _get_workspace_id(request)
        try:
            import csv
            import io
            path = _safe_path(workspace_id, filename)
            updated = [row[:] for row in sheet_data]  # deep copy

            # Separate delete_row ops and sort in REVERSE order to prevent index drift
            ops = result.get('operations', [])
            delete_ops = sorted(
                [op for op in ops if op.get('op') == 'delete_row'],
                key=lambda o: o.get('row', 0),
                reverse=True
            )
            other_ops = [op for op in ops if op.get('op') != 'delete_row']

            # Apply non-delete ops first
            for op in other_ops:
                op_type = op.get('op')
                if op_type == 'set_cell':
                    r, c = op['row'], op['col']
                    while len(updated) <= r:
                        updated.append([])
                    while len(updated[r]) <= c:
                        updated[r].append('')
                    updated[r][c] = op['value']
                elif op_type == 'insert_row':
                    after = op.get('after_row', len(updated) - 1)
                    updated.insert(after + 1, op.get('values', []))
                elif op_type == 'insert_col':
                    after = op.get('after_col', len(updated[0]) - 1 if updated and updated[0] else 0)
                    header = op.get('header', '')
                    values = op.get('values', [])
                    for i, row in enumerate(updated):
                        val = header if i == 0 else (values[i - 1] if i - 1 < len(values) else '')
                        row.insert(after + 1, val)
                elif op_type == 'delete_col':
                    c = op['col']
                    for row in updated:
                        if 0 <= c < len(row):
                            row.pop(c)

            # Apply delete_row ops in reverse order (highest index first)
            for op in delete_ops:
                r = op['row']
                if 0 <= r < len(updated):
                    updated.pop(r)

            buf = io.StringIO()
            writer = csv.writer(buf)
            for row in updated:
                writer.writerow(row)
            path.write_text(buf.getvalue(), encoding='utf-8')
        except Exception as e:
            save_warning = f'Operations applied in UI but file save failed: {str(e)}'
            print(f'[AI] File save error for {filename}: {e}')

    response_data = {'status': 'ok', **result}
    if save_warning:
        response_data['warning'] = save_warning
    return JsonResponse(response_data)
