"""
ai_tools.py - Sandboxed file operations the AI can invoke.

Every path is resolved relative to MEDIA_ROOT / 'ai_workspace' / <session_id>.
Path traversal attempts such as '../../manage.py' are blocked by _safe_path().
"""

import csv
import json
import os
from pathlib import Path

from django.conf import settings


def _env_int(name: str, default: int, minimum: int | None = None, maximum: int | None = None) -> int:
    try:
        value = int(os.environ.get(name, default))
    except (TypeError, ValueError):
        value = default
    if minimum is not None:
        value = max(minimum, value)
    if maximum is not None:
        value = min(maximum, value)
    return value


AI_MAX_FILE_WRITE_CHARS = _env_int('AI_MAX_FILE_WRITE_CHARS', 20000, minimum=1, maximum=200000)
AI_MAX_FILE_READ_CHARS = _env_int('AI_MAX_FILE_READ_CHARS', 6000, minimum=1, maximum=50000)
AI_MAX_CSV_ROWS = _env_int('AI_MAX_CSV_ROWS', 100, minimum=1, maximum=1000)
AI_MAX_LIST_FILES = _env_int('AI_MAX_LIST_FILES', 100, minimum=1, maximum=1000)


def _workspace_root(session_id: str) -> Path:
    """Return the per-session sandbox directory, creating it if needed."""
    root = Path(settings.MEDIA_ROOT) / 'ai_workspace' / session_id
    root.mkdir(parents=True, exist_ok=True)
    return root


def _safe_path(session_id: str, filename: str) -> Path:
    """Resolve filename inside the sandbox. Raises ValueError on escape."""
    if not filename or os.path.isabs(filename):
        raise ValueError(f"Access denied: invalid path '{filename}'.")

    root = _workspace_root(session_id)
    target = (root / filename).resolve()
    try:
        target.relative_to(root.resolve())
    except ValueError:
        raise ValueError(f"Access denied: path '{filename}' escapes the sandbox.")
    return target


def _truncate_text(text: str, limit: int) -> tuple[str, bool]:
    if len(text) <= limit:
        return text, False
    return text[:limit], True


def create_file(session_id: str, filename: str, content: str) -> dict:
    """Create or overwrite a file inside the user's AI workspace."""
    if len(content) > AI_MAX_FILE_WRITE_CHARS:
        return {'status': 'error', 'message': f'File content is too large. Limit is {AI_MAX_FILE_WRITE_CHARS} characters.'}

    path = _safe_path(session_id, filename)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding='utf-8')
    return {'status': 'ok', 'message': f'File "{filename}" created ({len(content)} chars).'}


def read_file(session_id: str, filename: str) -> dict:
    """Return capped text content of a file from the workspace."""
    path = _safe_path(session_id, filename)
    if not path.exists():
        return {'status': 'error', 'message': f'File "{filename}" does not exist.'}
    try:
        if path.name.lower().endswith('.xlsx'):
            import openpyxl
            import io
            import csv
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            output = io.StringIO()
            writer = csv.writer(output)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([str(cell) if cell is not None else "" for cell in row])
            text = output.getvalue()
        else:
            text = path.read_text(encoding='utf-8')
    except UnicodeDecodeError:
        return {'status': 'error', 'message': f'File "{filename}" is binary and cannot be read as text.'}
    except Exception as e:
        return {'status': 'error', 'message': f'Failed to read {filename}: {str(e)}'}

    text, truncated = _truncate_text(text, AI_MAX_FILE_READ_CHARS)
    return {'status': 'ok', 'filename': filename, 'content': text, 'truncated': truncated}


def edit_file(session_id: str, filename: str, search_text: str, replace_text: str) -> dict:
    """Replace the first occurrence of search_text with replace_text."""
    if len(replace_text) > AI_MAX_FILE_WRITE_CHARS:
        return {'status': 'error', 'message': f'Replacement text is too large. Limit is {AI_MAX_FILE_WRITE_CHARS} characters.'}

    path = _safe_path(session_id, filename)
    if not path.exists():
        return {'status': 'error', 'message': f'File "{filename}" does not exist.'}
    text = path.read_text(encoding='utf-8')
    if search_text not in text:
        return {'status': 'error', 'message': f'Search text not found in "{filename}".'}
    text = text.replace(search_text, replace_text, 1)
    path.write_text(text, encoding='utf-8')
    return {'status': 'ok', 'message': f'Replaced in "{filename}" successfully.'}


def append_to_file(session_id: str, filename: str, content: str) -> dict:
    """Append text to the end of a file."""
    if len(content) > AI_MAX_FILE_WRITE_CHARS:
        return {'status': 'error', 'message': f'Append content is too large. Limit is {AI_MAX_FILE_WRITE_CHARS} characters.'}

    path = _safe_path(session_id, filename)
    if not path.exists():
        return {'status': 'error', 'message': f'File "{filename}" does not exist.'}
    with open(path, 'a', encoding='utf-8') as f:
        f.write(content)
    return {'status': 'ok', 'message': f'Appended to "{filename}" successfully.'}


def list_files(session_id: str) -> dict:
    """List files in the workspace, capped to avoid huge tool responses."""
    root = _workspace_root(session_id)
    files = []
    truncated = False
    for p in root.rglob('*'):
        if p.is_file():
            rel = p.relative_to(root)
            files.append({'name': str(rel), 'size_bytes': p.stat().st_size})
            if len(files) >= AI_MAX_LIST_FILES:
                truncated = True
                break
    return {'status': 'ok', 'files': files, 'truncated': truncated}


def delete_file(session_id: str, filename: str) -> dict:
    """Delete a file from the workspace."""
    path = _safe_path(session_id, filename)
    if not path.exists():
        return {'status': 'error', 'message': f'File "{filename}" does not exist.'}
    path.unlink()
    return {'status': 'ok', 'message': f'File "{filename}" deleted.'}


def rename_file(session_id: str, old_filename: str, new_filename: str) -> dict:
    """Rename a file in the workspace."""
    old_path = _safe_path(session_id, old_filename)
    if not old_path.exists():
        return {'status': 'error', 'message': f'File "{old_filename}" does not exist.'}
        
    try:
        new_path = _safe_path(session_id, new_filename)
    except ValueError as e:
        return {'status': 'error', 'message': str(e)}

    if new_path.exists():
        return {'status': 'error', 'message': f'File "{new_filename}" already exists.'}
        
    old_path.rename(new_path)
    return {'status': 'ok', 'message': f'File renamed to "{new_filename}".'}


def add_row_to_csv(session_id: str, filename: str, row_data: str) -> dict:
    """Add a row to a CSV file. row_data is a JSON-encoded list of values."""
    path = _safe_path(session_id, filename)
    if not path.exists():
        return {'status': 'error', 'message': f'File "{filename}" does not exist.'}
    try:
        values = json.loads(row_data) if isinstance(row_data, str) else row_data
    except json.JSONDecodeError:
        return {'status': 'error', 'message': 'row_data must be a valid JSON array.'}
    with open(path, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(values)
    return {'status': 'ok', 'message': f'Row added to "{filename}".'}


def read_csv(session_id: str, filename: str) -> dict:
    """Read a CSV file and return capped rows."""
    path = _safe_path(session_id, filename)
    if not path.exists():
        return {'status': 'error', 'message': f'File "{filename}" does not exist.'}
    rows = []
    truncated = False
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for index, row in enumerate(reader):
            if index >= AI_MAX_CSV_ROWS:
                truncated = True
                break
            rows.append(row)
    return {'status': 'ok', 'filename': filename, 'rows': rows, 'truncated': truncated}


def create_excel_file(session_id: str, filename: str, sheet_data: str) -> dict:
    """Create an Excel (.xlsx) file. sheet_data is a JSON-encoded list of rows (which are lists of values)."""
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
        
    path = _safe_path(session_id, filename)
    if path.exists():
        return {'status': 'error', 'message': f'File "{filename}" already exists.'}
        
    try:
        data = json.loads(sheet_data) if isinstance(sheet_data, str) else sheet_data
        if not isinstance(data, list):
            return {'status': 'error', 'message': 'sheet_data must be a valid JSON 2D array.'}
    except json.JSONDecodeError:
        return {'status': 'error', 'message': 'sheet_data must be a valid JSON array.'}
        
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for row in data:
            if isinstance(row, list):
                ws.append(row)
        wb.save(path)
        return {'status': 'ok', 'message': f'Excel file "{filename}" created.'}
    except Exception as e:
        return {'status': 'error', 'message': f'Failed to create Excel file: {str(e)}'}


def append_to_excel_file(session_id: str, filename: str, row_data: str) -> dict:
    """Append a row of data to an existing Excel file."""
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
        
    path = _safe_path(session_id, filename)
    if not path.exists():
        return {'status': 'error', 'message': f'File "{filename}" does not exist. Use create_excel_file first.'}
        
    try:
        data = json.loads(row_data) if isinstance(row_data, str) else row_data
        if not isinstance(data, list):
            return {'status': 'error', 'message': 'row_data must be a valid JSON array.'}
    except json.JSONDecodeError:
        return {'status': 'error', 'message': 'row_data must be a valid JSON array.'}
        
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        ws.append(data)
        wb.save(path)
        return {'status': 'ok', 'message': f'Row appended to "{filename}".'}
    except Exception as e:
        return {'status': 'error', 'message': f'Failed to append to Excel file: {str(e)}'}


def read_excel_file(session_id: str, filename: str) -> dict:
    """Read an Excel file and return capped rows."""
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
        
    path = _safe_path(session_id, filename)
    if not path.exists():
        return {'status': 'error', 'message': f'File "{filename}" does not exist.'}
        
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = []
        truncated = False
        for index, row in enumerate(ws.iter_rows(values_only=True)):
            if index >= AI_MAX_CSV_ROWS:
                truncated = True
                break
            # Convert any non-serializable objects (like datetime) to strings
            clean_row = [str(cell) if cell is not None else "" for cell in row]
            rows.append(clean_row)
        return {'status': 'ok', 'filename': filename, 'rows': rows, 'truncated': truncated}
    except Exception as e:
        return {'status': 'error', 'message': f'Failed to read Excel file: {str(e)}'}


def execute_python(session_id: str, code: str) -> dict:
    """Execute a python script in the workspace directory. Returns stdout/stderr."""
    import subprocess
    import tempfile
    
    workspace = _workspace_root(session_id)
    if not workspace.exists():
        workspace.mkdir(parents=True)
        
    script_path = workspace / 'ai_temp_script.py'
    try:
        script_path.write_text(code, encoding='utf-8')
        
        # Run python script with timeout
        # Using the same python executable that the django app runs on
        import sys
        result = subprocess.run(
            [sys.executable, str(script_path)],
            cwd=str(workspace),
            capture_output=True,
            text=True,
            timeout=15
        )
        output = result.stdout
        if result.stderr:
            output += f"\n[STDERR]:\n{result.stderr}"
            
        return {
            'status': 'ok' if result.returncode == 0 else 'error',
            'output': output[:AI_MAX_FILE_READ_CHARS],
            'return_code': result.returncode
        }
    except subprocess.TimeoutExpired:
        return {'status': 'error', 'message': 'Script timed out after 15 seconds.'}
    except Exception as e:
        return {'status': 'error', 'message': str(e)}
    finally:
        if script_path.exists():
            script_path.unlink()


TOOL_DECLARATIONS = [
    {
        'name': 'create_file',
        'description': 'Create or overwrite a text file in the user workspace.',
        'parameters': {
            'type': 'object',
            'properties': {
                'filename': {'type': 'string', 'description': 'Name of the file, for example "report.txt" or "data.csv".'},
                'content': {'type': 'string', 'description': 'Full text content to write.'},
            },
            'required': ['filename', 'content'],
        },
    },
    {
        'name': 'read_file',
        'description': 'Read a text file from the user workspace and return capped content.',
        'parameters': {
            'type': 'object',
            'properties': {'filename': {'type': 'string', 'description': 'Name of the file to read.'}},
            'required': ['filename'],
        },
    },
    {
        'name': 'edit_file',
        'description': 'Edit a file by replacing the first occurrence of search_text with replace_text.',
        'parameters': {
            'type': 'object',
            'properties': {
                'filename': {'type': 'string', 'description': 'Name of the file to edit.'},
                'search_text': {'type': 'string', 'description': 'Exact text to find.'},
                'replace_text': {'type': 'string', 'description': 'Text to replace it with.'},
            },
            'required': ['filename', 'search_text', 'replace_text'],
        },
    },
    {
        'name': 'append_to_file',
        'description': 'Append text content to the end of an existing file.',
        'parameters': {
            'type': 'object',
            'properties': {
                'filename': {'type': 'string', 'description': 'Name of the file.'},
                'content': {'type': 'string', 'description': 'Text to append.'},
            },
            'required': ['filename', 'content'],
        },
    },
    {
        'name': 'list_files',
        'description': 'List files in the user workspace with their sizes.',
        'parameters': {'type': 'object', 'properties': {}, 'required': []},
    },
    {
        'name': 'delete_file',
        'description': 'Delete a file from the user workspace.',
        'parameters': {
            'type': 'object',
            'properties': {'filename': {'type': 'string', 'description': 'Name of the file to delete.'}},
            'required': ['filename'],
        },
    },
    {
        'name': 'add_row_to_csv',
        'description': 'Append a row of data to an existing CSV file.',
        'parameters': {
            'type': 'object',
            'properties': {
                'filename': {'type': 'string', 'description': 'Name of the CSV file.'},
                'row_data': {'type': 'string', 'description': 'A JSON array of values, for example ["Alice", 30, "Engineer"].'},
            },
            'required': ['filename', 'row_data'],
        },
    },
    {
        'name': 'read_csv',
        'description': 'Read a CSV file and return capped rows.',
        'parameters': {
            'type': 'object',
            'properties': {
                'filename': {'type': 'string', 'description': 'Name of the CSV file.'},
            },
            'required': ['filename'],
        },
    },
    {
        'name': 'create_excel_file',
        'description': 'Create an Excel (.xlsx) file with tabular data.',
        'parameters': {
            'type': 'object',
            'properties': {
                'filename': {'type': 'string', 'description': 'Name of the .xlsx file.'},
                'sheet_data': {'type': 'string', 'description': 'A JSON array containing the rows (e.g. [["Name", "Age"], ["Alice", 25], ["Bob", 30]]).'},
            },
            'required': ['filename', 'sheet_data'],
        },
    },
    {
        'name': 'append_to_excel_file',
        'description': 'Append a single row of data to an existing Excel (.xlsx) file. NEVER delete an Excel file to edit it, always append to it instead.',
        'parameters': {
            'type': 'object',
            'properties': {
                'filename': {'type': 'string', 'description': 'Name of the .xlsx file.'},
                'row_data': {'type': 'string', 'description': 'A JSON array containing the row values (e.g. ["Alice", 25, "Engineer"]).'},
            },
            'required': ['filename', 'row_data'],
        },
    },
    {
        'name': 'read_excel_file',
        'description': 'Read an existing Excel (.xlsx) file and return the rows as JSON. Use this to inspect the columns before appending.',
        'parameters': {
            'type': 'object',
            'properties': {
                'filename': {'type': 'string', 'description': 'Name of the .xlsx file.'},
            },
            'required': ['filename'],
        },
    },
    {
        'name': 'execute_python',
        'description': 'Write and execute a Python script in the user workspace. Excellent for using pandas/openpyxl to modify complex Excel or CSV files with 100% precision.',
        'parameters': {
            'type': 'object',
            'properties': {
                'code': {'type': 'string', 'description': 'The exact Python code to execute. You can use standard libraries + pandas, openpyxl, numpy. Data is stored in the current working directory.'},
            },
            'required': ['code'],
        },
    },
]


TOOL_MAP = {
    'create_file': create_file,
    'read_file': read_file,
    'edit_file': edit_file,
    'append_to_file': append_to_file,
    'list_files': list_files,
    'delete_file': delete_file,
    'rename_file': rename_file,
    'add_row_to_csv': add_row_to_csv,
    'read_csv': read_csv,
    'create_excel_file': create_excel_file,
    'append_to_excel_file': append_to_excel_file,
    'read_excel_file': read_excel_file,
    'execute_python': execute_python,
}
