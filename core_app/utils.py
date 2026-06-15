import csv
import os


def convert_pdf_to_docx(pdf_file, docx_file):
    from pdf2docx import Converter
    cv = Converter(pdf_file)
    cv.convert(docx_file, start=0, end=None)
    cv.close()


def convert_docx_to_pdf(docx_file, pdf_file):
    from docx2pdf import convert
    convert(docx_file, pdf_file)


def convert_excel_to_csv(input_file, csv_file):
    """Convert Excel (.xlsx, .xls) or Apple Numbers (.numbers) to CSV."""
    ext = os.path.splitext(input_file)[1].lower()

    if ext == '.numbers':
        _numbers_to_csv(input_file, csv_file)
    elif ext == '.xlsx':
        _xlsx_to_csv(input_file, csv_file)
    elif ext == '.xls':
        _xls_to_csv(input_file, csv_file)
    else:
        raise ValueError('Unsupported file type. Use .xlsx, .xls, or .numbers')


def _xlsx_to_csv(input_file, csv_file):
    from openpyxl import load_workbook
    wb = load_workbook(input_file, read_only=True, data_only=True)
    ws = wb.active
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(['' if cell is None else cell for cell in row])
    wb.close()


def _xls_to_csv(input_file, csv_file):
    import xlrd
    book = xlrd.open_workbook(input_file)
    sheet = book.sheet_by_index(0)
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row_idx in range(sheet.nrows):
            writer.writerow(sheet.row_values(row_idx))


def _numbers_to_csv(input_file, csv_file):
    try:
        from numbers_parser import Document
    except ImportError:
        raise ValueError('Apple Numbers support requires numbers-parser. Export as Excel first, or install the package.')

    doc = Document(input_file)
    sheets = doc.sheets
    if not sheets:
        raise ValueError('No sheets found in Numbers file.')

    table = sheets[0].tables[0]
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in table.iter_rows():
            writer.writerow([cell.value if cell.value is not None else '' for cell in row])
